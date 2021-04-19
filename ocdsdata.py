import base64
import csv
import datetime
import functools
import gzip
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import traceback
import zipfile
from collections import Counter, deque
from pathlib import Path
from textwrap import dedent

import boto3
import click
import lxml.html
import ocdsmerge
import openpyxl
import orjson
import requests
import sqlalchemy as sa
from codetiming import Timer
from fastavro import parse_schema, writer
from google.cloud import bigquery
from google.cloud.bigquery.dataset import AccessEntry
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from jsonref import JsonRef
from ocdsextensionregistry import ProfileBuilder
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from scrapy import signals
from scrapy.crawler import CrawlerProcess
from scrapy.spiderloader import SpiderLoader
from scrapy.utils.project import get_project_settings


this_path = Path(__file__).parent.absolute()
collect_path = str((this_path / "kingfisher-collect").absolute())


def _first_doc_line(function):
    return function.__doc__.split("\n")[0]


def _patched_schema(connection):
    extension_result = connection.execute(
        """
        SELECT
            extension
        FROM
            _package_data, jsonb_array_elements(package_data -> 'extensions') extension
        WHERE
            jsonb_typeof(package_data -> 'extensions') = 'array' group by 1;
    """
    )

    extensions = [row.extension for row in extension_result]

    try:
        builder = ProfileBuilder("1__1__4", extensions)
        patched_schema = builder.patched_release_schema()
        print(f'Exensions being used: {" ".join(extensions)}')
    except Exception:
        print('Warning: Extensions have not been built defaulting to standard schema')
        builder = ProfileBuilder("1__1__4", {})
        patched_schema = builder.patched_release_schema()

    return patched_schema


@functools.lru_cache(None)
def get_engine(schema=None, db_uri=None, pool_size=1):
    """Get SQLAlchemy engine

    Will cache engine if all arguments are the same so not expensive to call multiple times.


    Parameters
    ----------
    schema : string, optional
        Postgres schema that all queries will use. Defaults to using public schema.
    db_url : string, optional
        SQLAlchemy database connection string. Will defailt to using `DATABASE_URL` environment variable.
    pool_size : int
       SQLAlchemy connection pool size


    Returns
    -------
    sqlalchemy.Engine
        SQLAlchemy Engine object set up to query specified schema (or public schema)
    """

    if not db_uri:
        db_uri = os.environ["DATABASE_URL"]

    connect_args = {}
    if schema:
        connect_args = {"options": f"-csearch_path={schema}"}

    return sa.create_engine(db_uri, pool_size=pool_size, connect_args=connect_args)


def get_s3_bucket():
    """Get S3 bucket object

    Needs environment variables:

    `AWS_ACCESS_KEY_ID`,
    `AWS_S3_ENDPOINT_URL`,
    `AWS_SECRET_ACCESS_KEY`,
    `AWS_DEFAULT_REGION`,
    `AWS_S3_ENDPOINT_URL`

    Returns
    -------
    s3.Bucket
        s3.Bucket object to interact with S3

    """

    session = boto3.session.Session()
    if not os.environ.get("AWS_ACCESS_KEY_ID"):
        return

    s3 = session.resource("s3", endpoint_url=os.environ.get("AWS_S3_ENDPOINT_URL"))
    bucket = s3.Bucket(os.environ.get("AWS_S3_BUCKET"))
    return bucket


def get_drive_service():
    json_acct_info = orjson.loads(
        base64.b64decode(os.environ["GOOGLE_SERVICE_ACCOUNT"])
    )
    credentials = service_account.Credentials.from_service_account_info(
        json_acct_info
    )

    return build("drive", "v3", credentials=credentials)


def create_table(table, schema, sql, **params):
    """Create table under given schema by supplying SQL

    Parameters
    ----------
    table : string
        Postgres schema to use.
    schema : string
        Postgres schema to use.
    sql : string
        SQL to create table can be parametarized by SQLAlchemy parms that start with a `:` e.g `:param`.
    params : key (string), values (any)
        keys are params found in sql and values are the values to be replaced.
    """
    print(f"creating table {table}")
    t = Timer()
    t.start()
    engine = get_engine(schema)
    with engine.connect() as con:
        con.execute(
            sa.text(
                f"""DROP TABLE IF EXISTS {table};
                    CREATE TABLE {table}
                    AS
                    {sql};"""
            ),
            **params,
        )
    t.stop()


@click.group()
def cli():
    pass


def scraper_list():
    """List of scrapers from kingfisher collect.

    Returns
    -------
    list of str
        List of scrapers
    """
    os.chdir(collect_path)
    settings = get_project_settings()
    sl = SpiderLoader.from_settings(settings)
    return sl.list()


@cli.command(help=_first_doc_line(scraper_list))
def export_scrapers():
    click.echo(json.dumps(scraper_list()))


@cli.command()
@click.argument("name")
@click.argument("schema")
def import_scraper(name, schema):
    """Create all postgres tables for a scraper into a target schema."""
    create_schema(schema)
    scrape(name, schema)
    create_base_tables(schema, drop_scrape=False)
    compile_releases(schema)
    release_objects(schema)
    schema_analysis(schema)
    postgres_tables(schema, drop_release_objects=False)


@cli.command()
@click.argument("schema")
@click.argument("name")
@click.argument("date")
def export_all(name, schema, date):
    """Export data to all export soruces from schama given a date."""
    export_csv(schema, name, date)
    export_xlsx(schema, name, date)
    export_sqlite(schema, name, date)
    export_bigquery(schema, name, date)
    export_stats(schema, name, date)
    export_pgdump(schema, name, date)


def create_schema(schema):
    """Create Postgres Schema.

    Parameters
    ----------
    schema : string
        Postgres schema to create.
    """
    engine = get_engine()
    with engine.begin() as connection:
        connection.execute(
            f"""DROP SCHEMA IF EXISTS {schema} CASCADE;
                create schema {schema};"""
        )


@cli.command("create-schema", help=_first_doc_line(create_schema))
@click.argument("schema")
def _create_schema(schema):
    create_schema(schema)


def rename_schema(schema, new_schema):
    """Rename Postgres Schema.

    Parameters
    ----------
    schema : string
        Postgres schema to rename.
    new_schema : string
        New schema name.
    """
    engine = get_engine()
    drop_schema(new_schema)
    with engine.begin() as connection:
        connection.execute(f"""ALTER SCHEMA "{schema}" RENAME TO "{new_schema}";""")


@cli.command("rename-schema", help=_first_doc_line(rename_schema))
@click.argument("schema")
@click.argument("new_schema")
def _rename_schema(schema, new_schema):
    rename_schema(schema, new_schema)


def drop_schema(schema):
    """Drop Postgres Schema.

    Parameters
    ----------
    schema : string
        Postgres schema to drop.
    """
    engine = get_engine()
    with engine.begin() as connection:
        connection.execute(f"""DROP SCHEMA IF EXISTS {schema} CASCADE;""")


@cli.command("drop-schema", help=_first_doc_line(drop_schema))
@click.argument("schema")
def _drop_schema(schema):
    drop_schema(schema)


def scrape(name, schema):
    """Scrape data into postgres.

    Creates "_scrape_data" and "_job_info" tables.

    Parameters
    ----------
    name : string
        Name of scraper.
    schema : string
        Postgres schema to put scrape data in.
    """
    data_dir = this_path / "data" / schema
    shutil.rmtree(data_dir, ignore_errors=True)
    data_dir.mkdir(parents=True, exist_ok=True)
    csv_file_path = data_dir / "all.csv"

    engine = get_engine(schema)
    with engine.begin() as connection:
        connection.execute(
            """DROP TABLE IF EXISTS _scrape_data;
               DROP TABLE IF EXISTS _job_info;
               CREATE TABLE _scrape_data(id SERIAL, name TEXT, url TEXT, data_type TEXT, file_name TEXT,
                                         valid BOOLEAN, data JSONB, error_data TEXT);
               CREATE TABLE _job_info(name TEXT, info JSONB, logs TEXT);
            """
        )

    os.chdir(collect_path)
    settings = get_project_settings()
    settings.pop("FILES_STORE")
    settings["LOG_FILE"] = str(data_dir / "all.log")

    with gzip.open(str(csv_file_path), "wt", newline="") as csv_file:

        csv_writer = csv.writer(csv_file)

        count_data_types = Counter()

        def save_to_csv(item, spider):
            if "error" in item:
                print(item)
                return
            if "data" not in item:
                print(item)
                return
            try:
                if isinstance(item["data"], dict):
                    data = orjson.dumps(item["data"], default=str)
                    valid = "t"
                    error_data = ""
                else:
                    try:
                        orjson.loads(item["data"])
                        data = item["data"]
                        valid = "t"
                        error_data = ""
                    except orjson.JSONDecodeError:
                        valid = "f"
                        data = "{}"
                        error_data = item["data"]

                if isinstance(data, bytes):
                    data = data.decode(item["encoding"])
                data = data.replace(r"\u0000", "")

                count_data_types.update([item["data_type"]])

                csv_writer.writerow(
                    [
                        name,
                        item["url"],
                        item["data_type"],
                        item["file_name"],
                        valid,
                        data,
                        error_data,
                    ]
                )
            except Exception:
                traceback.print_exc()
                raise

        runner = CrawlerProcess(settings)
        crawler = runner.create_crawler(name)

        crawler.signals.connect(save_to_csv, signal=signals.item_scraped)

        runner.crawl(crawler)
        runner.start()

    info = crawler.stats.spider_stats[name]
    info["name"] = name
    info["data_types"] = dict(count_data_types)
    info_file = data_dir / "info.json"
    info_data = json.dumps(info, default=str)
    info_file.write_text(info_data)

    log_file = data_dir / "all.log"

    def tail(filename, n=10000):
        with open(filename) as file:
            return "".join(deque(file, n))

    with engine.begin() as connection, gzip.open(str(csv_file_path), "rt") as f:
        connection.execute(
            sa.text("INSERT INTO _job_info VALUES (:name, :info, :logs)"),
            name=name,
            info=info_data,
            logs=tail(log_file),
        )

        dbapi_conn = connection.connection
        copy_sql = "COPY _scrape_data (name, url, data_type, file_name, valid, data, error_data) FROM STDIN WITH CSV"
        cur = dbapi_conn.cursor()
        cur.copy_expert(copy_sql, f)

        result = connection.execute("SELECT count(*) FROM _scrape_data").first()

        print(f"{result['count']} files scraped")
        if result.count == 0:
            print("No data scraped!")
            sys.exit(1)

    shutil.rmtree(data_dir)


@cli.command("scrape", help=_first_doc_line(scrape))
@click.argument("name")
@click.argument("schema")
def _scrape(name, schema):
    scrape(name, schema)


def create_base_tables(schema, drop_scrape=True):
    """Create "_compiled_release" and "_package_data" tables"

    Parameters
    ----------
    schema : string
        Postgres schema where the "_scrape_data" table is.
    drop_scrape : boolean default True
        Drop the _scrape_data table when finished with it.
    """
    engine = get_engine(schema)

    package_data_sql = """
       SELECT
           id,
           data - 'releases' - 'records' package_data
       FROM
           _scrape_data
       WHERE
           data_type in ('release_package', 'record_package')
    """
    create_table("_package_data", schema, package_data_sql)

    engine.execute(
        """
        drop sequence IF EXISTS _generated_release_id;
        create sequence _generated_release_id;
        """
    )

    compiled_releases_sql = """
       SELECT
           min(nextval('_generated_release_id')) compiled_release_id,
           name,
           data_type,
           jsonb_agg(id) package_data_ids,
           coalesce(release ->> 'ocid', gen_random_uuid()::text) ocid,
           jsonb_agg(release) release_list,
           null rest_of_record,
           null compiled_release,
           null compile_error
       FROM
           _scrape_data,
           jsonb_path_query(data, '$.releases[*]') release
       WHERE
           data_type in ('release_package')

       GROUP BY name, data_type, coalesce(release ->> 'ocid', gen_random_uuid()::text)

       UNION ALL

       SELECT
           nextval('_generated_release_id'),
           name,
           data_type,
           jsonb_build_array(id),
           record ->> 'ocid',
           record -> 'releases',
           record - 'compiledRelease' rest_of_record,
           record -> 'compiledRelease' compiled_release,
           null compile_error
       FROM
           _scrape_data,
           jsonb_path_query(data, '$.records[*]') record
       WHERE
           data_type in ('record_package')
    """

    create_table("_compiled_releases", schema, compiled_releases_sql)

    if drop_scrape:
        engine.execute("DROP TABLE IF EXISTS _scrape_data")

    result = engine.execute("SELECT count(*) FROM _compiled_releases").first()

    print(f"{result['count']} compiled releases")
    if result.count == 0:
        print("No compiled releases!")
        sys.exit(1)


@cli.command("create-base-tables", help=_first_doc_line(create_base_tables))
@click.argument("schema")
def _create_base_tables(schema):
    create_base_tables(schema)


def compile_releases(schema):
    """Merge releases into a compiled_release.

    For release packages merge releases to _compiled_release.compiled_release column.

    Parameters
    ----------
    schema : string
        Postgres schema where the "_scrape_data" table is.
    """
    with tempfile.TemporaryDirectory() as tmpdirname:
        csv_file_path = tmpdirname + "/compiled_release.csv"

        engine = get_engine(schema)

        engine.execute(
            """
            DROP TABLE IF EXISTS _tmp_compiled_releases;
            CREATE TABLE _tmp_compiled_releases(compiled_release_id bigint, compiled_release JSONB, compile_error TEXT)
        """
        )

        patched_schema = _patched_schema(engine)

        merger = ocdsmerge.Merger(patched_schema)

        print("Making CSV file")
        with gzip.open(str(csv_file_path), "wt", newline="") as csv_file, engine.begin() as connection, Timer():
            connection = connection.execution_options(stream_results=True, max_row_buffer=1000)
            results = connection.execute(
                "SELECT compiled_release_id, release_list FROM _compiled_releases WHERE compiled_release is null"
            )
            csv_writer = csv.writer(csv_file)
            for num, result in enumerate(results):
                try:
                    compiled_release = merger.create_compiled_release(
                        result.release_list
                    )
                    error = ""
                except Exception as e:
                    compiled_release = {}
                    error = str(e)
                csv_writer.writerow(
                    [result.compiled_release_id, json.dumps(compiled_release), error]
                )

        print("Importing file")
        with engine.begin() as connection, Timer(), gzip.open(
            str(csv_file_path), "rt"
        ) as f:
            dbapi_conn = connection.connection
            copy_sql = "COPY _tmp_compiled_releases FROM STDIN WITH CSV"
            cur = dbapi_conn.cursor()
            cur.copy_expert(copy_sql, f)

        print("Updating table")
        with engine.begin() as connection, Timer():
            connection.execute(
                """UPDATE _compiled_releases cr
                   SET compiled_release = tmp.compiled_release,
                       compile_error = tmp.compile_error
                   FROM _tmp_compiled_releases tmp
                   WHERE tmp.compiled_release_id = cr.compiled_release_id"""
            )
            connection.execute(
                """
                DROP TABLE IF EXISTS _tmp_compiled_releases;
            """
            )


@cli.command("compile-releases", help=_first_doc_line(compile_releases))
@click.argument("schema")
def _compile_releases(schema):
    compile_releases(schema)


EMIT_OBJECT_PATHS = [
    ("planning",),
    ("tender",),
    ("contracts", "implementation"),
    ("buyer",),
    ("tender", "procuringEntity"),
]

PARTIES_PATHS = [
    "buyer",
    "awards_suppliers",
    "tender_procuringEntity",
    "tender_tenderers",
]


def flatten_object(obj, current_path=""):
    for key, value in list(obj.items()):
        if isinstance(value, dict):
            yield from flatten_object(value, f"{current_path}{key}_")
        else:
            yield f"{current_path}{key}", value


def traverse_object(obj, emit_object, full_path=tuple(), no_index_path=tuple()):
    for key, value in list(obj.items()):
        if isinstance(value, list) and value and isinstance(value[0], dict):
            for num, item in enumerate(value):
                if not isinstance(item, dict):
                    item = {"__error": "A non object is in array of objects"}
                yield from traverse_object(
                    item, True, full_path + (key, num), no_index_path + (key,)
                )
            obj.pop(key)
        elif isinstance(value, list):
            if not all(isinstance(item, str) for item in value):
                obj[key] = json.dumps(value)
        elif isinstance(value, dict):
            if no_index_path + (key,) in EMIT_OBJECT_PATHS:
                yield from traverse_object(
                    value, True, full_path + (key,), no_index_path + (key,)
                )
                obj.pop(key)
            else:
                yield from traverse_object(
                    value, False, full_path + (key,), no_index_path + (key,)
                )

    if obj and emit_object:
        yield obj, full_path, no_index_path


@functools.lru_cache(1000)
def path_info(full_path, no_index_path):
    all_paths = []
    for num, part in enumerate(full_path):
        if isinstance(part, int):
            all_paths.append(full_path[: num + 1])

    parent_paths = all_paths[:-1]
    path_key = all_paths[-1] if all_paths else []

    object_key = ".".join(str(key) for key in path_key)
    parent_keys_list = [
        ".".join(str(key) for key in parent_path) for parent_path in parent_paths
    ]
    parent_keys_no_index = [
        "_".join(str(key) for key in parent_path if not isinstance(key, int))
        for parent_path in parent_paths
    ]
    object_type = "_".join(str(key) for key in no_index_path) or "release"
    parent_keys = (dict(zip(parent_keys_no_index, parent_keys_list)),)
    return object_key, parent_keys_list, parent_keys_no_index, object_type, parent_keys


def create_rows(result):
    rows = []
    awards = {}
    parties = {}
    for object, full_path, no_index_path in traverse_object(result.compiled_release, 1):

        (
            object_key,
            parent_keys_list,
            parent_keys_no_index,
            object_type,
            parent_keys,
        ) = path_info(full_path, no_index_path)

        object[
            "_link"
        ] = f'{result.compiled_release_id}{"." if object_key else ""}{object_key}'
        object["_link_release"] = str(result.compiled_release_id)
        for no_index_path, full_path in zip(parent_keys_no_index, parent_keys_list):
            object[
                f"_link_{no_index_path}"
            ] = f"{result.compiled_release_id}.{full_path}"

        row = dict(
            compiled_release_id=result.compiled_release_id,
            object_key=object_key,
            parent_keys=parent_keys,
            object_type=object_type,
            object=object,
        )

        rows.append(row)
        if object_type == "awards":
            award_id = object.get("id")
            if award_id:
                awards[award_id] = object
        if object_type == "parties":
            parties_id = object.get("id")
            parties[parties_id] = object

    for row in rows:
        object = row["object"]
        if row["object_type"] in PARTIES_PATHS:
            parties_id = object.get("id")
            if parties_id:
                party = parties.get(parties_id)
                if party:
                    object["_link_party"] = party["_link"]
                    object["_party"] = {
                        key: value
                        for key, value in party.items()
                        if not key.startswith("_")
                    }
        if row["object_type"] == "contracts":
            award_id = object.get("awardID")
            if award_id:
                award = awards.get(award_id)
                if award:
                    object["_link_award"] = award["_link"]
                    object["_award"] = {
                        key: value
                        for key, value in award.items()
                        if not key.startswith("_")
                    }
        try:
            row["object"] = orjson.dumps(dict(flatten_object(object))).decode()
        except TypeError:
            # orjson more strict about ints
            row["object"] = json.dumps(dict(flatten_object(object)))

        row["parent_keys"] = orjson.dumps(row["parent_keys"]).decode()

    return [list(row.values()) for row in rows]


@cli.command("release-objects")
@click.argument("schema")
def _release_objects(schema):
    release_objects(schema)


def release_objects(schema):
    engine = get_engine(schema)
    engine.execute(
        """
        DROP TABLE IF EXISTS _release_objects;
        CREATE TABLE _release_objects(compiled_release_id bigint,
        object_key TEXT, parent_keys JSONB, object_type TEXT, object JSONB);
        """
    )

    with tempfile.TemporaryDirectory() as tmpdirname:
        with engine.begin() as connection, Timer():
            connection = connection.execution_options(stream_results=True, max_row_buffer=1000)
            results = connection.execute(
                "SELECT compiled_release_id, compiled_release FROM _compiled_releases"
            )
            paths_csv_file = tmpdirname + "/paths.csv"

            print("Making CSV file")
            with gzip.open(paths_csv_file, "wt", newline="") as csv_file, Timer():
                csv_writer = csv.writer(csv_file)
                for result in results:
                    csv_writer.writerows(create_rows(result))

        print("Uploading Data")
        with engine.begin() as connection, gzip.open(
            paths_csv_file, "rt"
        ) as f, Timer():
            dbapi_conn = connection.connection
            copy_sql = f"COPY {schema}._release_objects FROM STDIN WITH CSV"
            cur = dbapi_conn.cursor()
            cur.copy_expert(copy_sql, f)


def process_schema_object(path, current_name, flattened, obj):
    string_path = ("_".join(path)) or "release"

    properties = obj.get("properties", {})  # an object may have patternProperties only
    current_object = flattened.get(string_path)

    if current_object is None:
        current_object = {}
        flattened[string_path] = current_object

    for name, prop in list(properties.items()):
        prop_type = prop["type"]
        prop_info = dict(
            schema_type=prop["type"],
            description=prop.get("description"),
        )
        if prop_type == "object":
            if path + (name,) in EMIT_OBJECT_PATHS:
                flattened = process_schema_object(
                    path + (name,), tuple(), flattened, prop
                )
            else:
                flattened = process_schema_object(
                    path, current_name + (name,), flattened, prop
                )
        elif prop_type == "array":
            if "object" not in prop["items"]["type"]:
                current_object["_".join(current_name + (name,))] = prop_info
            else:
                flattened = process_schema_object(
                    path + current_name + (name,), tuple(), flattened, prop["items"]
                )
        else:
            current_object["_".join(current_name + (name,))] = prop_info

    return flattened


def link_info(link_name):
    name = link_name[6:]
    if not name:
        doc = "Link to this row that can be found in other rows"
    else:
        doc = f"Link to the {name} row that this row relates to"

    return {"name": link_name, "description": doc, "type": "string"}


@cli.command("schema-analysis")
@click.argument("schema")
def _schema_analysis(schema):
    schema_analysis(schema)


DATE_RE = r'^(\d{4})-(\d{2})-(\d{2})([T ](\d{2}):(\d{2}):(\d{2}(?:\.\d*)?)((-(\d{2}):(\d{2})|Z)?))?$'


def schema_analysis(schema):

    create_table(
        "_object_type_aggregate",
        schema,
        f"""SELECT
              object_type,
              each.key,
              CASE
                 WHEN jsonb_typeof(value) != 'string'
                     THEN jsonb_typeof(value)
                 WHEN (value ->> 0) ~ '{DATE_RE}'
                     THEN 'datetime'
                 ELSE 'string'
              END value_type,
              count(*)
           FROM
              _release_objects ro, jsonb_each(object) each
           GROUP BY 1,2,3;
        """,
    )

    create_table(
        "_object_type_fields",
        schema,
        """SELECT
              object_type,
              key,
              CASE WHEN
                  count(*) > 1
              THEN 'string'
              ELSE max(value_type) end value_type,
              SUM("count") AS "count"
           FROM
              _object_type_aggregate
           WHERE
              value_type != 'null'
           GROUP BY 1,2;
        """,
    )

    patched_schema = _patched_schema(get_engine(schema))

    schema_info = process_schema_object(
        tuple(), tuple(), {}, JsonRef.replace_refs(patched_schema)
    )

    with get_engine(schema).begin() as connection:
        result = connection.execute(
            """SELECT object_type, jsonb_object_agg(key, value_type) fields FROM _object_type_fields GROUP BY 1;"""
        )
        result_dict = {row.object_type: row.fields for row in result}

        object_type_order = ["release"]
        for key in schema_info:
            if key in result_dict:
                object_type_order.append(key)
        for key in result_dict:
            if key not in object_type_order:
                object_type_order.append(key)

        object_details = {}

        for object_type in object_type_order:
            fields = result_dict[object_type]

            details = [link_info("_link"), link_info("_link_release")]
            fields_added = set(["_link", "_link_release"])

            for field in fields:
                if field.startswith("_link_") and field not in fields_added:
                    details.append(link_info(field))
                    fields_added.add(field)

            schema_object_detials = schema_info.get(object_type, {})

            for schema_field, field_info in schema_object_detials.items():
                if schema_field not in fields:
                    continue
                detail = {"name": schema_field, "type": fields[schema_field]}
                detail.update(field_info)
                details.append(detail)
                fields_added.add(schema_field)

            for field in sorted(fields):
                if field in fields_added:
                    continue
                details.append(
                    {
                        "name": field,
                        "description": "No Docs as not in OCDS",
                        "type": fields[field],
                    }
                )

            object_details[object_type] = details

        connection.execute(
            """
            DROP TABLE IF EXISTS _object_details;
            CREATE TABLE _object_details(id SERIAL, object_type text, object_details JSONB);
        """
        )

        for object_type, object_details in object_details.items():
            connection.execute(
                sa.text(
                    "insert into _object_details(object_type, object_details) values (:object_type, :object_details)"
                ),
                object_type=object_type,
                object_details=json.dumps(object_details),
            )


def create_field_sql(object_details, sqlite=False):
    fields = []
    lowered_fields = set()
    fields_with_type = []
    for num, item in enumerate(object_details):
        name = item["name"]

        if sqlite and name.lower() in lowered_fields:
            name = f'{name}_{num}'

        type = item["type"]
        if type == "number":
            field = f'"{name}" numeric'
        elif type == "array":
            field = f'"{name}" JSONB'
        elif type == "boolean":
            field = f'"{name}" boolean'
        elif type == "datetime":
            field = f'"{name}" timestamp'
        else:
            field = f'"{name}" TEXT'

        lowered_fields.add(name.lower())
        fields.append(f'"{name}"')
        fields_with_type.append(field)

    return ", ".join(fields), ", ".join(fields_with_type)


@cli.command("postgres-tables")
@click.argument("schema")
def _postgres_tables(schema):
    postgres_tables(schema)


def postgres_tables(schema, drop_release_objects=True):
    with get_engine(schema).begin() as connection:
        result = list(
            connection.execute(
                "SELECT object_type, object_details FROM _object_details order by id"
            )
        )

    for object_type, object_details in result:
        field_sql, as_sql = create_field_sql(object_details)
        table_sql = f"""
           SELECT {field_sql}
           FROM _release_objects, jsonb_to_record(object) AS ({as_sql})
           WHERE object_type = :object_type
        """
        create_table(object_type, schema, table_sql, object_type=object_type)

    if drop_release_objects:
        with get_engine(schema).begin() as connection:
            connection.execute("DROP TABLE IF EXISTS _release_objects")


def generate_object_type_rows(object_detials_results):
    yield ["Sheet Name", "Name", "Docs", "Type", "Schema Type"]

    for object_type, object_details in object_detials_results:
        for field in object_details:
            yield [
                object_type,
                field.get("name"),
                field.get("description"),
                field.get("type"),
                str(field.get("schema_type")),
            ]


@cli.command("export-csv")
@click.argument("schema")
@click.argument("name")
@click.argument("date")
def _export_csv(schema, name, date):
    export_csv(schema, name, date)


def export_csv(schema, name, date):
    with tempfile.TemporaryDirectory() as tmpdirname, get_engine(
        schema
    ).begin() as connection:
        with zipfile.ZipFile(f"{tmpdirname}/output.zip", "w") as zip_file:
            result = list(
                connection.execute(
                    "SELECT object_type, object_details FROM _object_details order by id"
                )
            )
            for object_type, object_details in result:
                with open(f"{tmpdirname}/{object_type}.csv", "wb") as out:
                    dbapi_conn = connection.connection
                    copy_sql = f'COPY "{object_type.lower()}" TO STDOUT WITH CSV HEADER'
                    cur = dbapi_conn.cursor()
                    cur.copy_expert(copy_sql, out)

                zip_file.write(
                    f"{tmpdirname}/{object_type}.csv",
                    arcname=f"{name}/{object_type}.csv",
                )

            with open(f"{tmpdirname}/fields.csv", "w") as csv_file:
                csv_writer = csv.writer(csv_file)
                csv_writer.writerows(generate_object_type_rows(result))
                zip_file.write(f"{tmpdirname}/fields.csv", arcname=f"{name}/fields.csv")

        bucket = get_s3_bucket()
        if bucket:
            object = bucket.Object(f"{name}/ocdsdata_{name}_csv.zip")
            object.upload_file(
                f"{tmpdirname}/output.zip", ExtraArgs={"ACL": "public-read"}
            )
            metadata_object = bucket.Object(f"{name}/metadata/csv_upload_dates/{date}")
            metadata_object.put(ACL="public-read", Body=b"")


def shorten_sheet_name(name):
    replacements = [
        ("tender_procuringEntity", "procuringEntity"),
        ("additionalIdentifiers", "ids"),
        ("additionalClassifications", "class"),
        ("awards_suppliers", "suppliers"),
        ("contracts_implementation_transactions", "transactions"),
        ("contracts_implementation", "implementation"),
        ("document", "doc"),
    ]
    for field, replacement in replacements:
        name = name.replace(field, replacement)

    return name[:31]


def generate_spreadsheet_rows(result, object_details):

    for row in result:
        data = row.data
        line = []
        for field in object_details:
            value = data[field["name"]] or ""
            if isinstance(value, list):
                value = ", ".join(value)
            if field["type"] == "string":
                value = str(value)
                new_value = ILLEGAL_CHARACTERS_RE.sub("", value)
                if new_value != value:
                    print(
                        f"""Character(s) in '{value}' are not allowed in a spreadsheet cell.
                            Those character(s) will be removed"""
                    )
                value = new_value
            line.append(value)
        yield line


@cli.command("export-xlsx")
@click.argument("schema")
@click.argument("name")
@click.argument("date")
def _export_xlsx(schema, name, date):
    export_xlsx(schema, name, date)


def export_xlsx(schema, name, date):
    with tempfile.TemporaryDirectory() as tmpdirname, get_engine(
        schema
    ).begin() as connection:

        largest_row_count = (
            connection.execute(
                "SELECT max(count) as largest_row_count FROM _object_type_fields"
            )
            .first()
            .largest_row_count
        )
        print(f"Larges row count {largest_row_count}")
        if largest_row_count > 100000:
            print("More than 100000 rows not creating xlsx")
            return

        object_details_result = list(
            connection.execute(
                "SELECT object_type, object_details FROM _object_details order by id"
            )
        )

        workbook = openpyxl.Workbook(write_only=True)

        worksheet = workbook.create_sheet()
        worksheet.title = "Field Information"
        for row in generate_object_type_rows(object_details_result):
            worksheet.append(row)

        for object_type, object_details in object_details_result:
            result = connection.execute(
                sa.text(
                    f'SELECT to_json("{object_type.lower()}") AS data FROM "{object_type.lower()}"'
                )
            )

            worksheet = workbook.create_sheet()
            worksheet.title = shorten_sheet_name(object_type)
            sheet_header = [item["name"] for item in object_details]
            worksheet.append(sheet_header)

            for row in generate_spreadsheet_rows(result, object_details):
                worksheet.append(row)

        workbook.save(f"{tmpdirname}/data.xlsx")

        bucket = get_s3_bucket()
        if bucket:
            object = bucket.Object(f"{name}/ocdsdata_{name}.xlsx")
            object.upload_file(
                f"{tmpdirname}/data.xlsx", ExtraArgs={"ACL": "public-read"}
            )
            metadata_object = bucket.Object(f"{name}/metadata/xlsx_upload_dates/{date}")
            metadata_object.put(ACL="public-read", Body=b"")


name_allowed_pattern = re.compile(r"[\W]+")


def create_avro_schema(object_type, object_details):
    fields = []
    schema = {"type": "record", "name": object_type, "fields": fields}
    for item in object_details:
        type = item["type"]
        if type == "number":
            type = "double"
        if type == "datetime":
            type = "string"

        field = {
            "name": name_allowed_pattern.sub("", item["name"]),
            "type": [type, "null"],
            "doc": item.get("description"),
        }

        if type == "array":
            field["type"] = [
                {"type": "array", "items": "string", "default": []},
                "null",
            ]

        fields.append(field)

    return schema


def generate_avro_records(result, object_details):

    cast_to_string = set(
        [field["name"] for field in object_details if field["type"] == "string"]
    )

    for row in result:
        new_object = {}
        for key, value in row.object.items():
            new_object[name_allowed_pattern.sub("", key)] = (
                str(value) if key in cast_to_string else value
            )
        yield new_object


@cli.command("export-bigquery")
@click.argument("schema")
@click.argument("name")
@click.argument("date")
def _export_bigquery(schema, name, date):
    export_bigquery(schema, name, date)


def export_bigquery(schema, name, date):

    json_acct_info = orjson.loads(
        base64.b64decode(os.environ["GOOGLE_SERVICE_ACCOUNT"])
    )

    credentials = service_account.Credentials.from_service_account_info(json_acct_info)

    client = bigquery.Client(credentials=credentials)

    with tempfile.TemporaryDirectory() as tmpdirname, get_engine(
        schema
    ).begin() as connection:
        dataset_id = f"{client.project}.{name}"
        client.delete_dataset(dataset_id, delete_contents=True, not_found_ok=True)
        dataset = bigquery.Dataset(dataset_id)
        dataset.location = "EU"

        dataset = client.create_dataset(dataset, timeout=30)

        access_entries = list(dataset.access_entries)
        access_entries.append(
            AccessEntry("READER", "specialGroup", "allAuthenticatedUsers")
        )
        dataset.access_entries = access_entries

        dataset = client.update_dataset(dataset, ["access_entries"])

        result = connection.execute(
            "SELECT object_type, object_details FROM _object_details order by id"
        )
        for object_type, object_details in list(result):
            print(f"loading {object_type}")
            result = connection.execute(
                sa.text(
                    f'SELECT to_jsonb("{object_type.lower()}") AS object FROM "{object_type.lower()}"'
                )
            )
            schema = create_avro_schema(object_type, object_details)

            with open(f"{tmpdirname}/{object_type}.avro", "wb") as out:
                writer(
                    out,
                    parse_schema(schema),
                    generate_avro_records(result, object_details),
                    validator=True,
                    codec="deflate",
                )

            table_id = f"{client.project}.{name}.{object_type}"

            job_config = bigquery.LoadJobConfig(
                source_format=bigquery.SourceFormat.AVRO
            )

            with open(f"{tmpdirname}/{object_type}.avro", "rb") as source_file:
                client.load_table_from_file(
                    source_file, table_id, job_config=job_config, size=None, timeout=5
                )

            bucket = get_s3_bucket()
            if bucket:
                object = bucket.Object(
                    f"{name}/avro/ocdsdata_{name}_{object_type}.avro"
                )
                object.upload_file(
                    f"{tmpdirname}/{object_type}.avro", ExtraArgs={"ACL": "public-read"}
                )
                metadata_object = bucket.Object(
                    f"{name}/metadata/avro_upload_dates/{date}"
                )
                metadata_object.put(ACL="public-read", Body=b"")


@cli.command("export-sqlite")
@click.argument("schema")
@click.argument("name")
@click.argument("date")
def _export_sqlite(schema, name, date):
    export_sqlite(schema, name, date)


def export_sqlite(schema, name, date):
    with tempfile.TemporaryDirectory() as tmpdirname, get_engine(
        schema
    ).begin() as connection:
        result = list(
            connection.execute(
                "SELECT object_type, object_details FROM _object_details order by id"
            )
        )
        for object_type, object_details in result:
            print(f"importing table {object_type}")
            with open(f"{tmpdirname}/{object_type}.csv", "wb") as out:
                dbapi_conn = connection.connection
                copy_sql = f'COPY "{object_type.lower()}" TO STDOUT WITH CSV'
                cur = dbapi_conn.cursor()
                cur.copy_expert(copy_sql, out)

            _, field_def = create_field_sql(object_details, sqlite=True)
            import_sql = f"""
            .mode csv
            CREATE TABLE "{object_type}" ({field_def}) ;
            .import {tmpdirname}/{object_type}.csv "{object_type}" """

            subprocess.run(
                ["sqlite3", f"{tmpdirname}/{name}.sqlite"],
                input=dedent(import_sql),
                text=True,
                check=True,
            )

            os.remove(f"{tmpdirname}/{object_type}.csv")

        with open(f"{tmpdirname}/fields.csv", "w") as csv_file:
            csv_writer = csv.writer(csv_file)
            csv_writer.writerows(generate_object_type_rows(result))

        import_sql = f"""
            .mode csv
            .import {tmpdirname}/fields.csv "fields" """

        subprocess.run(
            ["sqlite3", f"{tmpdirname}/{name}.sqlite"],
            input=dedent(import_sql),
            text=True,
            check=True,
        )

        subprocess.run(
            ["gzip", "-k", "-9", f"{tmpdirname}/{name}.sqlite"],
            check=True
        )

        bucket = get_s3_bucket()
        if bucket:
            object = bucket.Object(f"{name}/ocdsdata_{name}.sqlite")
            object.upload_file(
                f"{tmpdirname}/{name}.sqlite", ExtraArgs={"ACL": "public-read"}
            )
            object = bucket.Object(f"{name}/ocdsdata_{name}.sqlite.gz")
            object.upload_file(
                f"{tmpdirname}/{name}.sqlite.gz", ExtraArgs={"ACL": "public-read"}
            )
            metadata_object = bucket.Object(
                f"{name}/metadata/sqlite_upload_dates/{date}"
            )
            metadata_object.put(ACL="public-read", Body=b"")


@cli.command("export-stats")
@click.argument("schema")
@click.argument("name")
@click.argument("date")
def _export_stats(schema, name, date):
    export_stats(schema, name, date)


def export_stats(schema, name, date):
    bucket = get_s3_bucket()

    with get_engine(schema).begin() as connection:
        result = connection.execute(
            "SELECT to_json(_job_info) AS job_info FROM _job_info"
        )
        job_info = orjson.dumps(result.first().job_info)
        job_info_object = bucket.Object(f"{name}/metadata/job_info/{date}.json")
        job_info_object.put(ACL="public-read", Body=job_info)

        result = connection.execute(
            "SELECT to_json(_object_type_fields) AS data FROM _object_type_fields"
        )
        field_info = orjson.dumps([row.data for row in result])
        field_info_object = bucket.Object(f"{name}/metadata/field_info/{date}.json")
        field_info_object.put(ACL="public-read", Body=field_info)

        result = connection.execute(
            "SELECT to_json(_object_details) AS data FROM _object_details"
        )
        field_info = orjson.dumps([row.data for row in result])
        field_info_object = bucket.Object(f"{name}/metadata/field_types/{date}.json")
        field_info_object.put(ACL="public-read", Body=field_info)


@cli.command("export-pgdump")
@click.argument("schema")
@click.argument("name")
@click.argument("date")
def _export_pgdump(schema, name, date):
    export_pgdump(schema, name, date)


def export_pgdump(schema, name, date):
    bucket = get_s3_bucket()

    with tempfile.TemporaryDirectory() as tmpdirname:
        subprocess.run(
            [
                "pg_dump",
                "--no-owner",
                "-f",
                tmpdirname + "/dump.sql",
                "-n",
                schema,
                "-F",
                "c",
                os.environ["DATABASE_URL"],
            ]
        )

        pg_dump_object = bucket.Object(f"{name}/ocdsdata_{name}.pg_dump")
        pg_dump_object.upload_file(
            f"{tmpdirname}/dump.sql", ExtraArgs={"ACL": "public-read"}
        )
        metadata_object = bucket.Object(f"{name}/metadata/pg_dump_upload_dates/{date}")
        metadata_object.put(ACL="public-read", Body=b"")


def _make_table_markdown(schema, info):
    all_field_types = requests.get(info['field_types']['latest']).json()

    field_type_strings = {}

    for field_types in all_field_types:
        fields_str = ", ".join([f'{object_detail["name"]}' for object_detail in  field_types["object_details"]])
        if len(fields_str) > 200:
            fields_str = fields_str[:200] + '...'

        field_type_strings[field_types["object_type"]] = fields_str

    lines = ['| Table | Row Count | Fields |']
    lines.append('|-|-|-|')
    for table, count in info.get('table_stats', {}).items():
        lines.append(f'| [{table}](https://ocds-downloads.opendata.coop/source/{schema}#table-{table}) | {count} | {field_type_strings[table]} |')

    return json.dumps("\n".join(lines))[1:-1]

FOLDER_ID="1wNkeGqdDP3wQR4xe8yyFreOez9U47Vz8"

@cli.command("make-notebook")
@click.argument("schema")
def _make_notebook(schema):
    bucket = get_s3_bucket()
    stats_object = bucket.Object("metadata/stats.json")
    stats = json.load(stats_object.get()["Body"])
    for schema_in_stats, info in stats.items():
        if schema_in_stats == schema:
            make_notebook(schema, info)
            break


def make_notebook(schema, info, folder_id=FOLDER_ID):
    bucket = get_s3_bucket()

    with tempfile.TemporaryDirectory() as tmpdirname, open(
        this_path / "notebook/template.ipynb"
    ) as template:
        template_text = template.read()
        template_text = template_text.replace("zambia", schema)
        template_text = template_text.replace(
            "Zambia", schema.replace("_", " ").capitalize()
        )

        template_text = template_text.replace('{{tables}}', _make_table_markdown(schema, info))

        with open(tmpdirname + "/notebook.ipynb", "w+") as output:
            output.write(template_text)

        drive_service = get_drive_service()

        folder_id = folder_id
        file_metadata = {"name": f"{schema}.ipynb", "parents": [folder_id]}
        media = MediaFileUpload(
            tmpdirname + "/notebook.ipynb",
            mimetype="application/vnd.google.colaboratory",
        )
        file = (
            drive_service.files()
            .create(body=file_metadata, media_body=media, fields="id")
            .execute()
        )

        pg_dump_object = bucket.Object(f"{schema}/ocdsdata_{schema}_notebook.json")
        pg_dump_object.put(ACL="public-read", Body=orjson.dumps(file))


def make_notebooks(stats=None, folder_id=FOLDER_ID):
    for schema, info in stats.items():
        if info.get('sqlite_gz'):
            make_notebook(schema, info, folder_id)


def parse_collect_docs_scraper_info():


    docs_url = "https://kingfisher-collect.readthedocs.io/en/latest/spiders.html"

    content = requests.get(docs_url).content

    tree = lxml.html.fromstring(content)

    spiders_section = tree.cssselect("#spiders")[0]

    sections = spiders_section.cssselect(".section")

    scrapers = {}

    for section in sections:
        heading = section.cssselect("h2")[0].text

        dls = [dl for dl in section.iterchildren("dl")]
        if not dls:
            continue

        pres = section.cssselect("div pre")

        assert len(dls) == len(pres)

        for dl, pre in zip(dls, pres):
            text = pre.text_content().strip()
            assert "scrapy crawl" in text
            scraper = text.split(" ")[-1]

            crawler_id = dl.cssselect("dt")[0].attrib["id"]

            inner_dl = dl.cssselect("dd dl")
            extra_info = {}
            if len(inner_dl):
                dt_text = [
                    dt.text_content().strip() for dt in inner_dl[0].iterchildren("dt")
                ]
                dd_text = [
                    dd.text_content().strip() for dd in inner_dl[0].iterchildren("dd")
                ]
                extra_info = dict(zip(dt_text, dd_text))
            scrapers[scraper] = {
                "category": heading,
                "extra_info": extra_info,
                "docs_link": f"{docs_url}#{crawler_id}",
            }

    return scrapers


@cli.command("collect-stats")
def _collect_stats():
    collect_stats()


def collect_stats():

    out = {}

    scraper_info = parse_collect_docs_scraper_info()

    for scraper in scraper_list():
        out[scraper] = {
            "csv": {},
            "xlsx": {},
            "sqlite": {},
            "sqlite_gz": {},
            "pg_dump": {},
            "avro": {"files": {}},
            "big_query": {},
            "notebookIdFile": "",
            "field_info": {},
            "field_types": {},
            "table_stats": {},
            "job_info": {},
            "scraper_info": scraper_info.get(scraper, {}),
        }

    bucket = get_s3_bucket()

    bucket_url = f"{bucket.meta.client.meta.endpoint_url}/{bucket.name}"

    for item in sorted(bucket.objects.all(), key=lambda x: x.key.split("/")[-1]):
        item_url = f"{bucket_url}/{item.key}"
        parts = item.key.split("/")
        scraper = parts[0]

        if scraper not in out:
            continue

        file_name = parts[-1]

        if file_name.endswith("csv.zip"):
            out[scraper]["csv"].update(file_name=file_name, url=item_url)
        if file_name.endswith(".sqlite"):
            out[scraper]["sqlite"].update(file_name=file_name, url=item_url)
        if file_name.endswith(".sqlite.gz"):
            out[scraper]["sqlite_gz"].update(file_name=file_name, url=item_url)
        if file_name.endswith("pg_dump"):
            out[scraper]["pg_dump"].update(file_name=file_name, url=item_url)
        if file_name.endswith("xlsx"):
            out[scraper]["xlsx"].update(file_name=file_name, url=item_url)
        if file_name.endswith("avro"):
            obj = re.sub(f"^ocdsdata_{scraper}_", "", file_name[:-5])
            out[scraper]["avro"]["files"][obj] = item_url
            out[scraper]["big_query"].update(
                url=f"https://console.cloud.google.com/bigquery?project=ocdsdata&p=ocdsdata&d={scraper}&page=dataset"
            )
        if file_name.endswith("_notebook.json"):
            out[scraper]["notebookIdFile"] = item_url

        if parts[1] not in ("metadata", "metatdata"):
            continue

        if "upload_dates" in parts[2]:
            file_type = parts[2].replace("_upload_dates", "")
            out[scraper][file_type]["latest_date"] = file_name

        if "field_info" in parts[2]:
            out[scraper]["field_info"]["latest"] = item_url
            out[scraper]["field_info"]["latest_item"] = item
            out[scraper]["field_info"][file_name[:-5]] = item_url
            out[scraper]["field_info"]["latest_date"] = file_name[:-5]

        if "field_types" in parts[2]:
            out[scraper]["field_types"]["latest"] = item_url
            out[scraper]["field_types"]["latest_item"] = item
            out[scraper]["field_types"][file_name[:-5]] = item_url
            out[scraper]["field_types"]["latest_date"] = file_name[:-5]

        if "job_info" in parts[2]:
            out[scraper]["job_info"]["latest"] = item_url
            out[scraper]["job_info"]["latest_item"] = item
            out[scraper]["job_info"][file_name[:-5]] = item_url
            out[scraper]["job_info"]["latest_date"] = file_name[:-5]

    for scraper, data in out.items():
        latest_field_info_item = data["field_info"].pop("latest_item", None)
        latest_field_types_item = data["field_types"].pop("latest_item", None)
        if latest_field_info_item and latest_field_types_item:
            field_types_data = json.load(latest_field_types_item.get()["Body"])
            object_type_order = {field_types['object_type']: num for num, field_types in enumerate(field_types_data)}
            field_info_data = json.load(latest_field_info_item.get()["Body"])
            for item in sorted(field_info_data, key = lambda x: object_type_order[x['object_type']]):
                if item["key"] == "_link":
                    data["table_stats"][item["object_type"]] = item["count"]

        latest_job_info_item = data["job_info"].pop("latest_item", None)
        if latest_job_info_item:
            job_info_data = json.load(latest_job_info_item.get()["Body"])
            data["job_info"]["latest_info"] = job_info_data["info"]

    drive_service = get_drive_service()

    file_metadata = {
        'name': str(datetime.datetime.utcnow())[:19],
        'mimeType': 'application/vnd.google-apps.folder',
        "parents": [FOLDER_ID]
    }
    file = drive_service.files().create(body=file_metadata,
                                        fields='id').execute()

    make_notebooks(out, file.get('id'))

    stats_object = bucket.Object("metadata/stats.json")
    stats_object.put(ACL="public-read", Body=orjson.dumps(out))


if __name__ == "__main__":
    cli()
